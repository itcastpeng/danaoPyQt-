from PyQt5.QtWidgets import *
from PyQt5.QtWebEngineWidgets import *
from PyQt5.QtWebChannel import *
from PyQt5.QtCore import *
from win32 import win32api
from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import QDesktopWidget
from openpyxl.styles import Font, Alignment
from openpyxl import Workbook
from tkinter import *
from tkinter.filedialog import askdirectory
from time import sleep
from PyQt5 import sip
from mian.my_db import database_create_data
from mian.repeater_timing import timing_task
import multiprocessing
from mian.threading_task_pc.threading_task import shoulu_func, fugai_func
import sqlite3, os, json, time, datetime, tkinter.messagebox, threading, sys, requests, queue
from mian import settings
from mian.backend import insert_databse
import subprocess


class Danao_Inter_Action(QObject):
    def __init__(self):
        super(Danao_Inter_Action, self).__init__()
        self.process = multiprocessing.Process(target=timing_task.run)
        self.process.start()
        self.zhongdianci_select_task_detail = ''
        self.fugai_huoqu_id_fanhui_detail = ''
        self.dangqian_chaxunshoulu_time = ''
        self.dangqian_chaxunfugai_time = ''
        self.huoqu_shoulu_time_stamp = ''
        self.zhongdianci_page_detail = ''
        self.huoqu_fugai_time_stamp = ''
        self.huoqu_task_id_detail = ''
        self.shoulu_chaxun_page = ''
        self.zhongdianci_update = ''
        self.fugai_chaxun_page = ''
        self.fugai_chongfu_num = '0'
        self.shoulu_chongfu = '0'
        self.tiaoshu = '10'
        self.panduan = '0'
        self.process_fugai = multiprocessing.Process()
        self.process_shoulu = multiprocessing.Process()
        self.fugai_number = 0

    # 占位 助手
    def zhanwei_zhushou(self):
        pass

    # 登录 - 返回登陆参数
    def get_Loginvalue(self):
        # 查询数据库 返回用户信息
        sql = """select message from Login_message;"""
        objs = database_create_data.operDB(sql, 'select')
        if objs['data']:
            # print('objs====================> ',objs)
            return objs['data'][0][0]

    # 登录 - 获取登录参数 保存数据库
    def set_Loginvalue(self, data):
        print('获得登录参数---->: %s' % data)
        # 查询数据库 有数据更新 没数据创建
        sql = """select * from Login_message;"""
        objs = database_create_data.operDB(sql, 'select')
        sql_data = ''
        sql_two = ''
        for sql_data in objs['data']:
            sql_data = sql_data
        if sql_data:
            sql_two = """update Login_message set message='{data}' where id=1""".format(data=data)
        else:
            sql_two = """insert into Login_message values (1,'{data}')""".format(data=data)
        database_create_data.operDB(sql_two, 'update or insert')

    # 重点词监控 - 获取任务列表数据
    def get_zhongdianci_create_task_list_value(self):
        sql = """select * from task_List;"""
        objs = database_create_data.operDB(sql, 'select')
        data_list = []
        baifenbi = 0
        if objs['data']:
            for obj in objs['data']:
                id = obj[0]
                sql_two = """select count(id) from task_Detail where tid = '{}';""".format(id)
                objs_two = database_create_data.operDB(sql_two, 'select')
                count_obj = int(objs_two['data'][0][0])
                sql_three = """select count(id) from task_Detail where tid = '{}' and is_perform = '0';""".format(id)
                objs_three = database_create_data.operDB(sql_three, 'select')
                if objs_three['data']:
                    wancheng_obj = int(objs_three['data'][0][0])
                    if count_obj != 0:
                        baifenbi = wancheng_obj / count_obj * 100
                        if wancheng_obj == count_obj:
                            sql = """update task_List set task_status='1', zhixing = '0' where id='{}'""".format(id)
                            database_create_data.operDB(sql, 'update')
                    qiyong_status = '未启用'
                    if obj[1]:
                        qiyong_status = '已启用'
                    zhixing = False
                    if obj[8]:
                        zhixing = True
                    """  task_status         任务状态
                                             1 未查询
                                             2 查询中
                                             3 已完成"""
                    task_status = '未查询'
                    if obj[5] == 0:
                        task_status = '查询中'
                    if obj[5] == 1:
                        task_status = '已完成'
                    data_list.append({
                        "id": id,
                        "qiyong_status": qiyong_status,
                        "task_name": obj[2],
                        "task_jindu": obj[3],
                        "task_start_time": obj[4],
                        "task_status": task_status,
                        "search_engine": obj[6].split(','),
                        "mohupipei": obj[7],
                        "zhixing": obj[8],
                        "next_datetime": obj[9],
                        "keywords": obj[10],
                        'task_jindu': int(baifenbi),
                    })
        return json.dumps(data_list)

    # 重点词监控 - 增加任务列表
    def set_zhongdianci_create_value(self, data):
        if type(data) == str:
            """search_engine
              需要查询的搜索引擎
                 1 百度
                 2 搜狗
                 3 360
                 4 手机百度
                 5 手机搜狗
                 6 手机360
                 7 神马"""
            json_data = json.loads(data)
            # print('json_data ------------------- > ', json_data )
            qiyong_status = json_data['qiyong_status']
            task_name = json_data['task_name']
            task_jindu = json_data['task_jindu']
            task_start_time = json_data['task_start_time']
            search_engine = ','.join(json_data['search_engine'])
            mohupipei = json_data['mohupipei']
            keywords = json_data['keywords']
            zhixing = json_data['zhixing']
            if qiyong_status:
                qiyong_status = 1
            else:
                qiyong_status = 0
            if zhixing:
                zhixing = 1
            else:
                zhixing = 0
            now_datetime_date = datetime.date.today()
            now_date_datetime = datetime.datetime.today().strftime('%H:%M:%S')
            str_now_date = str(now_date_datetime)
            now_date = '1900-01-01' + ' ' + str_now_date
            # 作比较  转成 datetime 格式
            task_start_time = datetime.datetime.strptime(task_start_time, '%H:%M:%S')
            now_date = datetime.datetime.strptime(now_date, '%Y-%m-%d %H:%M:%S')
            next_datetime = ''
            if now_date > task_start_time:
                # '当前时间小于 开始时间'
                task_start_time = task_start_time.strftime('%H:%M:%S')
                #  如果当前时间小于 任务开始时间 那么 加一天
                add_one_day = (now_datetime_date + datetime.timedelta(days=1)).strftime('%Y-%m-%d')
                next_datetime = str(add_one_day) + ' ' + task_start_time
            else:
                # print('当前时间大于开始时间')
                task_start_time = task_start_time.strftime('%H:%M:%S')
                next_datetime = str(now_datetime_date) + ' ' + task_start_time
            values = (
                qiyong_status, task_name, task_jindu, task_start_time,
                search_engine, mohupipei, zhixing, next_datetime
            )
            sql = """insert into Task_List (qiyong_status, task_name, task_jindu, task_start_time, search_engine, mohupipei, zhixing, next_datetime) values {values};""".format(
                values=values)
            database_create_data.operDB(sql, 'insert')
            shibiecanshu = 'zhongdianci'
            insert_databse.insert_into(data, 0, shibiecanshu)

    # 重点词监控 - 查询修改前数据 获取修改任务列表id
    def set_zhongdianci_update_task_list_value(self, data):
        print('获取修改任务列表id ----------- > ', data)
        self.zhongdianci_update = data

    # 重点词监控 - 查询修改前数据 返回该id任务列表数据
    def get_zhongdianci_update_task_list_data_value(self):
        if self.zhongdianci_update:
            sql = """select * from Task_List where id = {};""".format(int(self.zhongdianci_update))
            # print(o_id, task_name, task_status, task_start_time, qiyong_status, search_engine, task_jindu, zhixing)
            objs = database_create_data.operDB(sql, 'select')
            data_list = {}
            for obj in objs['data']:
                qiyong_status = '未启用'
                if obj[1]:
                    qiyong_status = '已启用'

                zhixing = False
                if obj[8]:
                    zhixing = True
                """  task_status         任务状态
                                         1 未查询
                                         2 查询中
                                         3 已完成"""
                task_status = '未查询'
                if obj[5] == 2:
                    task_status = '查询中'
                if obj[5] == 3:
                    task_status = '已完成'
                search_engine_list = []
                search_engine_list.append(obj[6])
                data_list = {
                    "id": obj[0],
                    "qiyong_status": obj[1],
                    "task_name": obj[2],
                    "task_jindu": obj[3],
                    "task_start_time": obj[4],
                    "task_status": obj[5],
                    "search_engine": search_engine_list,
                    "mohupipei": obj[7],
                    "zhixing": obj[8],
                    "next_datetime": obj[9],
                    "keywords": obj[10]
                }
            # print('返回原数据 id为{}的数据-------------> '.format(self.zhongdianci_update), data_list)
            return json.dumps(data_list)

    # 重点词监护 - 修改任务列表数据
    def set_zhongdianci_update_data_value(self, update_data):
        if self.zhongdianci_update and update_data:
            """{
            "id":1,
            "qiyong_status":1,
            "task_name":"测试任务=678678",
            "task_jindu":0,
            "task_start_time":"20:18:10",
            "search_engine":["1"],
            "mohupipei":"测试条件",
            "keywords":"1"
            }"""
            now_date_datetime = datetime.datetime.today().strftime('%H:%M:%S')
            now_datetime_date = datetime.date.today()
            str_now_date = str(now_date_datetime)
            now_date = '1900-01-01' + ' ' + str_now_date
            if type(update_data) == str:
                json_update_data = json.loads(update_data)
                id = json_update_data['id']
                task_name = json_update_data['task_name']
                task_start_time = json_update_data['task_start_time']
                qiyong_status = json_update_data['qiyong_status']
                if qiyong_status:
                    qiyong_status = '1'
                else:
                    qiyong_status = '0'
                task_start_time_date = datetime.datetime.strptime(task_start_time, '%H:%M:%S')
                now_date = datetime.datetime.strptime(now_date, '%Y-%m-%d %H:%M:%S')
                if now_date < task_start_time_date:
                    task_start_time_date = task_start_time_date.strftime('%H:%M:%S')
                    # 如果当前时间小于 任务开始时间 那么 加一天
                    add_one_day = (now_datetime_date + datetime.timedelta(days=1)).strftime('%Y-%m-%d')
                    next_datetime = str(add_one_day) + ' ' + task_start_time_date
                else:
                    task_start_time_date = task_start_time_date.strftime('%H:%M:%S')
                    next_datetime = str(now_datetime_date) + ' ' + task_start_time_date
                # print(next_datetime)
                task_start_time_date = str(task_start_time_date)
                sql = """update Task_List set task_name='{task_name}', task_start_time='{task_start_time}', next_datetime='{next_datetime}', qiyong_status='{qiyong_status}' where id={id};""".format(
                    task_name=task_name,
                    task_start_time=task_start_time_date,
                    next_datetime=next_datetime,
                    id=id,
                    qiyong_status=qiyong_status)
                database_create_data.operDB(sql, 'update')

    # 重点词监护 - 清空为该id的任务 - 的详情数据
    def set_zhongdianci_select_id_task_detail_value(self, select_task_detail):
        if select_task_detail:
            sql = """select id from task_Detail where tid = {}""".format(select_task_detail)
            objs = database_create_data.operDB(sql, 'select')
            if objs['data']:
                for obj in objs['data']:
                    sql_two = """delete from task_Detail_Data where tid={}""".format(obj[0])
                    database_create_data.operDB(sql_two, 'delete')
                sql_three = """delete from task_Detail where tid = {};""".format(select_task_detail)
                database_create_data.operDB(sql_three, 'delete')

    # 重点词监护 - 获取任务id
    def set_zhongdianci_select_id_select_task_detail_value(self, data):
        self.zhongdianci_select_task_detail = data

    # 重点词监护 - 查询该任务的详情
    def get_zhongdianci_select_id_select_task_detail_value(self):
        data_list = []
        if self.zhongdianci_select_task_detail:
            print('----------------')
            sql = """select * from task_Detail where tid={};""".format(self.zhongdianci_select_task_detail)
            objs = database_create_data.operDB(sql, 'select')
            print('执行到这了========')
            for obj in objs['data']:
                data_list.append({
                    "id": obj[0],
                    "tid": obj[1],
                    "search_engine": obj[2],
                    "lianjie": obj[3],
                    "keywords": obj[4],
                    "mohupipei": obj[5],
                    "create_time": obj[6],
                })
        return str(data_list)

    # 重点词监护 - 立即监控 获取需要执行的任务id 调用定时器
    def set_pc_task_value(self, datas):
        self.huoqu_shoulu_time_stamp = int(time.time())
        self.dangqian_chaxunshoulu_time = datetime.datetime.today().strftime('%Y-%m-%d %H-%M-%S')
        json_data = json.loads(datas)
        lijijiankong = multiprocessing.Process(target=timing_task.lijijiankong, args=(json_data,))
        lijijiankong.start()

    # 重点词监护 - 删除单个或多个 任务
    def delete_or_batch_delete_task(self, delete_id):
        json_id = json.loads(delete_id)
        for data in json_id:
            sql = """select id from task_Detail where tid = {}""".format(data)
            objs = database_create_data.operDB(sql, 'select')
            for obj in objs['data']:
                sql_two = """delete from task_Detail_Data where tid = {}""".format(obj[0])
                database_create_data.operDB(sql_two, 'delete')
            sql_three = """delete from task_Detail where tid = {};""".format(data)
            database_create_data.operDB(sql_three, 'delete')
            sql_four = """delete from task_List where id ={}""".format(data)
            database_create_data.operDB(sql_four, 'delete')

    # 重点词监护 - 获取展示详情获取子任务id
    def set_task_id_view_The_subtasks_task(self, data):
        self.huoqu_task_id_detail = data

    # 重点词监护 - 重点词监护分页
    def set_huoqu_page_detail_value(self, data):
        self.zhongdianci_page_detail = data

    # 重点词监护 - 展示详情任务子任务
    def get_view_The_subtasks_detail(self):
        if self.huoqu_task_id_detail:
            sql_page = ''
            sql_count = """select b.task_name,count(a.id) from task_Detail as a, task_List as b where a.tid = '{}' and a.tid=b.id """.format(
                self.huoqu_task_id_detail)
            if self.zhongdianci_page_detail:
                if int(self.zhongdianci_page_detail) == 1:
                    start_page = 0
                else:
                    start_page = (int(self.zhongdianci_page_detail) - 1) * 10
                #  分页                                                                           开始行数         条数
                sql_page = """select * from task_Detail where tid = {huoqu_task_id_detail} limit '{start_page}', '{tiaoshu}';""".format(
                    huoqu_task_id_detail=self.huoqu_task_id_detail,
                    start_page=start_page,
                    tiaoshu=int(self.tiaoshu)
                )
            else:
                sql_page = """select * from task_Detail where tid = {} limit 10;""".format(self.huoqu_task_id_detail)
                # sql = """select * from task_Detail as A, task_Detail_Data as B where A.id=B.tid and A.tid={} limit 10;""".format(self.huoqu_task_id_detail)
            data_list = []
            headers_list = []
            exit_data_list = []
            sql_count = database_create_data.operDB(sql_count, 'select')
            objs = database_create_data.operDB(sql_page, 'select')
            task_name = sql_count['data'][0][0]
            count = sql_count['data'][0][1]
            if objs:
                for obj in objs['data']:
                    sql_two = """select create_time, paiming, is_shoulu from task_Detail_Data where tid = {} order by create_time desc limit 3""".format(
                        obj[0])
                    objs_two = database_create_data.operDB(sql_two, 'select')
                    data_detail_list = []
                    sanci_chaxun = {}
                    if objs_two:
                        is_shoulu = ''
                        for detaildata_obj in objs_two['data']:
                            if detaildata_obj[2] == 1:
                                is_shoulu = '是'
                            else:
                                is_shoulu = '否'
                            detail_create = detaildata_obj[0]
                            detail_paiming = detaildata_obj[1]
                            detail_shoulu = is_shoulu
                            if detail_create not in headers_list:
                                headers_list.append(detail_create)
                            sanci_chaxun[detail_create] = {
                                'detail_create': detail_create,
                                'shoulu': detail_shoulu,
                                'paiming': detail_paiming
                            }
                    data_list.append({
                        'id': obj[0],
                        'search_engine': obj[2],
                        'lianjie': obj[3],
                        'keywords': obj[4],
                        'mohupipei': obj[5],
                        'sanci_chaxun': sanci_chaxun,
                    })
                exit_data_list = {
                    'task_name': task_name,
                    'count_page': count,
                    'data_list': data_list,
                    'headers_list': headers_list
                }
            return json.dumps(exit_data_list)

    # 重点词监护 - 导出 excl 功能
    def set_save_select_results_task_excel_daochu(self, tid_data):
        if tid_data:
            now_date = datetime.datetime.today().strftime('%Y-%m-%d %H-%M-%S')
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=1, value="重点关键词")
            ws.cell(row=2, column=1, value="任务名称:")
            ws.cell(row=3, column=1, value="关键词")
            ws.cell(row=3, column=2, value="网址")
            ws.cell(row=3, column=3, value="搜索引擎")
            ws.cell(row=4, column=4, value="排名")
            ws.cell(row=4, column=5, value="排名")
            ws.cell(row=4, column=6, value="排名")
            ft1 = Font(name='宋体', size=22)
            a1 = ws['A1']
            a1.font = ft1

            # 合并单元格        开始行      结束行       哪列做改变       占几列
            ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=6)
            ws.merge_cells(start_row=3, end_row=4, start_column=1, end_column=1)
            ws.merge_cells(start_row=3, end_row=4, start_column=2, end_column=2)
            ws.merge_cells(start_row=3, end_row=4, start_column=3, end_column=3)

            # print('设置列宽')
            ws.column_dimensions['A'].width = 50
            ws.column_dimensions['B'].width = 45
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 30
            ws.column_dimensions['E'].width = 42
            ws.column_dimensions['F'].width = 39
            ws.column_dimensions['G'].width = 39
            ws.column_dimensions['H'].width = 39

            # print('设置行高')
            ws.row_dimensions[1].height = 28
            ws.row_dimensions[2].height = 38
            ws.row_dimensions[3].height = 28
            ws.row_dimensions[4].height = 20

            # print('文本居中')
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws['A2'].alignment = Alignment(horizontal='right', vertical='center')
            ws['B2'].alignment = Alignment(horizontal='left', vertical='center')
            ws['C3'].alignment = Alignment(horizontal='center', vertical='center')
            ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
            ws['B3'].alignment = Alignment(horizontal='center', vertical='center')
            ws['A4'].alignment = Alignment(horizontal='center', vertical='center')

            sql = """select task_name from task_List where id = {};""".format(tid_data)
            objs = database_create_data.operDB(sql, 'select')
            task_name = objs['data'][0][0]
            sql_two = """select * from task_Detail where tid = {};""".format(tid_data)
            objs_two = database_create_data.operDB(sql_two, 'select')
            row = 5
            for obj in objs_two['data']:
                tid = obj[0]
                if str(obj[2]) == '1':
                    yinqing = '百度'
                elif  str(obj[2]) == '4':
                    yinqing = '手机百度'
                elif str(obj[2]) == '3':
                    yinqing = '360'
                elif str(obj[2]) == '6':
                    yinqing = '手机360'
                else:
                    yinqing = ''
                sql = 'select * from task_Detail_Data where tid = {} order by create_time desc limit 3  ;'.format(tid)
                objs = database_create_data.operDB(sql, 'select')
                if objs:
                    column_p = 4
                    for obj_data in objs['data']:
                        create_time = obj_data[4]
                        paiming = obj_data[1]
                        ws.cell(row=3, column=column_p, value="{create_time}".format(create_time=create_time))
                        ws.cell(row=row, column=column_p, value="{paiming}".format(paiming=paiming))
                        column_p += 1
                ws.cell(row=2, column=2, value="任务名")
                ws.cell(row=row, column=1, value="{keywords}".format(keywords=obj[4]))
                ws.cell(row=row, column=3, value="{search_engine}".format(search_engine=yinqing))
                ws.cell(row=row, column=2, value="{lianjie}".format(lianjie=obj[3]))
                row += 1
            root = Tk()
            root.withdraw()  # 隐藏
            root.iconbitmap('./128.ico')
            dirname = askdirectory(parent=root, initialdir="/", title='选择导出路径 !')
            if dirname:
                if dirname == 'C:/':
                    tkinter.messagebox.showerror('错误', '请选择路径 !')
                    # tkinter.messagebox.showwarning('警告', '请选择路径 !')
                else:
                    file_name = dirname.replace('\\', '/') + '/' + '{}.xlsx'.format(task_name + '_' + now_date)
                    wb.save(file_name)
                    # print(file_name)
                    tkinter.messagebox.showinfo('提示', '生成完毕 !')
            else:
                print('点击取消')
            root.destroy()  # 销毁

    # 收录查询 - 筛选链接 调用多线程 保存数据库
    def set_shoulu_select_get_list_value(self, data):
        delete_sql = """delete from shoulu_Linshi_List;"""
        database_create_data.operDB(delete_sql, 'delete')
        self.huoqu_shoulu_time_stamp = int(time.time())
        self.dangqian_chaxunshoulu_time = datetime.datetime.today().strftime('%Y-%m-%d %H-%M-%S')
        data_dict = json.loads(data)
        shibiecanshu = 'shoulu'
        insert_databse.insert_into(data, self.huoqu_shoulu_time_stamp, shibiecanshu)
        data_url_list = data_dict['editor_content'].replace('\r\n', '').strip().split('http')
        len_url_data = (len(data_url_list) * len(data_dict['searchEngineModel']))
        set_url_data = (len(set(data_url_list)) * len(data_dict['searchEngineModel']))
        self.shoulu_chongfu = len_url_data - set_url_data
        sleep(1)
        self.process_shoulu = multiprocessing.Process(target=shoulu_func,
            args=(self.huoqu_shoulu_time_stamp, set_url_data))
        self.process_shoulu.start()

    # 收录查询 - 收录分页处理
    def set_shoulu_chauxn_page_value(self, data):
        self.shoulu_chaxun_page = data

    # 收录查询 - 查询数据库 展示
    def get_shoulu_zhanshi_list_value(self):
        if self.huoqu_shoulu_time_stamp:
            data_list = []
            exit_dict = {}
            count_obj = 0
            if self.shoulu_chaxun_page:
                shoulu_page = json.loads(self.shoulu_chaxun_page)
                count_sql = """select count(id) from shoulu_Linshi_List where time_stamp = '{time_stamp}';""".format(
                    time_stamp=self.huoqu_shoulu_time_stamp)
                count_objs = database_create_data.operDB(count_sql, 'select')
                if count_objs['data']:
                    count_obj = count_objs['data'][0][0]  # 查询总数
                if int(shoulu_page['current_page']) == 1:
                    start_page = 0
                else:
                    start_page = (int(shoulu_page['current_page']) - 1) * 10
                limit_sql = """select * from shoulu_Linshi_List where time_stamp='{time_stamp}' limit '{start_page}', '{stop_page}';""".format(
                    time_stamp=self.huoqu_shoulu_time_stamp,
                    start_page=start_page,
                    stop_page=int(self.tiaoshu))
                objs = database_create_data.operDB(limit_sql, 'select')
                shoulushu_sql = """select count(id) from shoulu_Linshi_List where time_stamp = '{time_stamp}' and is_shoulu='1'; """.format(
                    time_stamp=self.huoqu_shoulu_time_stamp)
                shoulushu_objs = database_create_data.operDB(shoulushu_sql, 'select')
                shoulushu = shoulushu_objs['data'][0][0]
                shoululv = 0
                if shoulushu and count_obj != 0:
                    shoululv = int((shoulushu / count_obj) * 100)

                yiwancheng_obj = '0'
                yiwancheng_sql = """select count(id) from shoulu_Linshi_List where time_stamp='{time_stamp}' and is_zhixing='1';""".format(
                    time_stamp=self.huoqu_shoulu_time_stamp)
                yiwancheng_objs = database_create_data.operDB(yiwancheng_sql, 'select')
                if yiwancheng_objs['data']:
                    yiwancheng_obj = yiwancheng_objs['data'][0][0]
                    # 判断 是否全部执行完毕
                whether_complete = False
                if yiwancheng_obj == count_obj:
                    whether_complete = True
                for obj in objs['data']:
                    if obj[2] == '0':
                        shoulu_status = False
                    elif obj[2] == '1':
                        shoulu_status = True
                    else:
                        shoulu_status = ''
                    if str(obj[5]) == '1':
                        yinqing = '百度'
                    elif str(obj[5]) == '4':
                        yinqing = '手机百度'
                    elif str(obj[5]) == '3':
                        yinqing = '360'
                    elif str(obj[5]) == '6':
                        yinqing = '手机360'
                    else:
                        yinqing = ''
                    data_list.append({
                        'website': obj[1],
                        'shoulu_status': shoulu_status,
                        'title': obj[4],
                        'search_engine': yinqing,
                        'kuaizhao_date': obj[6],
                        'statusCode': obj[7],
                    })
                print('----收录---------收录----> ',yiwancheng_obj)
                exit_dict = {
                    'data': data_list,
                    'shoulushu': shoulushu,
                    'shoululv': shoululv,
                    'chongfu_num': self.shoulu_chongfu,
                    'whether_complete': whether_complete,  # 全部完成 传True
                    'count_obj': count_obj,  # 数据总数
                    'yiwancheng_obj': yiwancheng_obj  # 当前完成数量
                }
            return json.dumps(exit_dict)

    # 收录查询 - 查询数据库 导出excel表格
    def set_shoulu_save_select_result_value(self, data):
        if self.huoqu_shoulu_time_stamp:
            now_date = datetime.datetime.today().strftime('%Y-%m-%d %H-%M-%S')
            sql = """select url,is_shoulu,title,search,kuaizhao_time from shoulu_Linshi_List where {time_stamp};""".format(
                time_stamp=self.huoqu_shoulu_time_stamp)
            # time_stamp=huoqu_shoulu_time_stamp)
            objs = database_create_data.operDB(sql, 'select')
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=1, value="收录查询")
            ws.cell(row=2, column=4, value="查询时间:")
            ws.cell(row=6, column=1, value="标题")
            ws.cell(row=6, column=2, value="网址")
            ws.cell(row=6, column=3, value="搜索引擎")
            ws.cell(row=6, column=4, value="收录")
            ws.cell(row=6, column=5, value="快照日期")
            ft1 = Font(name='宋体', size=22)
            a1 = ws['A1']
            a1.font = ft1

            # 合并单元格        开始行      结束行       用哪列          占用哪列
            ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=6)
            ws.merge_cells(start_row=2, end_row=5, start_column=1, end_column=1)
            ws.merge_cells(start_row=2, end_row=5, start_column=2, end_column=2)
            ws.merge_cells(start_row=2, end_row=5, start_column=3, end_column=3)
            ws.merge_cells(start_row=2, end_row=5, start_column=4, end_column=4)
            ws.merge_cells(start_row=2, end_row=5, start_column=5, end_column=5)

            # print('设置列宽')
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 30
            ws.column_dimensions['E'].width = 30
            ws.column_dimensions['F'].width = 30

            # print('设置行高')
            ws.row_dimensions[1].height = 28

            # print('文本居中')
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws['D2'].alignment = Alignment(horizontal='center', vertical='center')
            ws['E2'].alignment = Alignment(horizontal='center', vertical='center')
            row = 7
            chaxun_time = self.dangqian_chaxunshoulu_time
            ws.cell(row=2, column=5, value="{chaxun_time}".format(chaxun_time=chaxun_time))
            for obj in objs['data']:
                is_shoulu = ''
                search = ''
                if obj[1] == '1':
                    is_shoulu = '已收录'
                else:
                    is_shoulu = '未收录'
                if str(obj[3]) == '1':
                    yinqing = '百度'
                elif  str(obj[3]) == '4':
                    yinqing = '手机百度'
                elif str(obj[3]) == '3':
                    yinqing = '360'
                elif str(obj[3]) == '6':
                    yinqing = '手机360'
                else:
                    yinqing = ''
                ws.cell(row=row, column=1, value="{title}".format(title=obj[2]))
                ws.cell(row=row, column=2, value="{url}".format(url=obj[0]))
                ws.cell(row=row, column=3, value="{search}".format(search=yinqing))
                ws.cell(row=row, column=4, value="{is_shoulu}".format(is_shoulu=is_shoulu))
                ws.cell(row=row, column=5, value="{kuaizhao_time}".format(kuaizhao_time=obj[4]))
                row += 1

            root = Tk()
            root.withdraw()
            dirname = askdirectory(parent=root, initialdir="/", title='Pick a directory')
            task_name = '收录查询'
            if dirname:
                if dirname == 'C:/':
                    tkinter.messagebox.showerror('错误', '请选择路径 !')
                else:
                    file_name = dirname.replace('\\', '/') + '/' + '{}.xls' \
                                                                   'x'.format(task_name + '_' + now_date)
                    wb.save(file_name)
                    tkinter.messagebox.showinfo('提示', '生成完毕 !')
            else:
                print('点击取消')
            root.destroy()  # 销毁

    # 收录查询 - 退出清空查询结果
    def set_shoulu_delete_all_values(self, data):
        if data == 'delete_all_shoulu':
            process_shoulu = os.popen('taskkill /PID %s /F' % self.process_shoulu.pid)
            subprocess.Popen(process_shoulu, shell=True)
            delete_sql = """delete from shoulu_Linshi_List;"""
            database_create_data.operDB(delete_sql, 'delete')

    # 覆盖查询 - 筛选查询条件 调用多线程 保存到数据库
    def set_fugai_select_get_list_value(self, data):
        self.fugai_number = 0
        delete_sql = """delete from fugai_Linshi_List;"""
        database_create_data.operDB(delete_sql, 'delete')
        json_data = json.loads(data)
        data_keyword_list = json_data['editor_content'].strip().split('\n')
        len_Keyword_data = len(data_keyword_list * len(json_data['searchEngineModel']))
        set_keyword_data = (len(set(data_keyword_list)) * len(json_data['searchEngineModel']))
        self.fugai_chongfu_num = len_Keyword_data - set_keyword_data
        self.huoqu_fugai_time_stamp = int(time.time())
        self.dangqian_chaxunfugai_time = datetime.datetime.today().strftime('%Y-%m-%d %H-%M-%S')
        shibiecanshu = 'fugai'
        if json_data['searchEngineModel'] and json_data['fugai_tiaojian']:
            insert_databse.insert_into(json_data, self.huoqu_fugai_time_stamp, shibiecanshu)
            sleep(1)
            self.process_fugai = multiprocessing.Process(target=fugai_func,
                args=(self.huoqu_fugai_time_stamp, set_keyword_data))
            self.process_fugai.start()

    # 覆盖查询 - 接收页码数
    def set_fugai_chaxun_xiangqing(self, data):
        data_dict = json.loads(data)
        self.fugai_chaxun_page = data_dict['currentPage']

    # 覆盖查询 - 获取时间戳 返回展示所有数据
    def get_fugai_zhanshi_list_value(self):
        start_page = ''
        data_list = []
        paiminglv = 0
        fugailv = 0
        paiming_num = 0
        yiwancheng_obj = 0
        whether_complete = False
        if self.huoqu_fugai_time_stamp:
            if not self.fugai_number:
                # 查询总数
                count_sql = """select count(id) from fugai_Linshi_List where time_stamp='{time_stamp}';""".format(
                    time_stamp=self.huoqu_fugai_time_stamp)
                count_objs = database_create_data.operDB(count_sql, 'select')
                if count_objs['data']:
                    self.fugai_number = int(count_objs['data'][0][0])
            # 已查询数量
            yiwancheng_sql = """select count(id) from fugai_Linshi_List where time_stamp='{time_stamp}' and is_zhixing = '1';""".format(
                time_stamp=self.huoqu_fugai_time_stamp)
            yiwancheng_objs = database_create_data.operDB(yiwancheng_sql, 'select')
            if yiwancheng_objs['data']:
                yiwancheng_obj = yiwancheng_objs['data'][0][0]

            if self.fugai_chaxun_page:
                if self.fugai_chaxun_page == 1:
                    start_page = 0
                else:
                    start_page = (self.fugai_chaxun_page - 1) * 10
                objs_sql = """select * from fugai_Linshi_List limit '{start_page}', '{tiaoshu}';""".format(
                    start_page=start_page,
                    tiaoshu=int(self.tiaoshu)
                )
                select_objs = database_create_data.operDB(objs_sql, 'select')
                rank_info = ''
                if select_objs['data']:
                    for detail_obj in select_objs['data']:
                        otherData = []
                        rank_num = 0
                        if detail_obj[1]:
                            rank_info = detail_obj[1]
                            rank_num = len(rank_info.split(','))
                        if detail_obj[1] == '0' or None:
                            rank_info = '_'
                            rank_num = 0
                        if str(detail_obj[2]) == '1':
                            yinqing = '百度'
                        elif str(detail_obj[2]) == '4':
                            yinqing = '手机百度'
                        elif str(detail_obj[2]) == '3':
                            yinqing = '360'
                        elif str(detail_obj[2]) == '6':
                            yinqing = '手机360'
                        else:
                            yinqing = ''
                        detail_obj_json = ''
                        if detail_obj[8]:
                            detail_obj_json = json.loads(detail_obj[8])
                        data_list.append({
                            'id': detail_obj[0],
                            'keyword': detail_obj[10],
                            'rank_info': rank_info,  # 排名情况  为空为查询中  无排名为-
                            'search_engine': yinqing,  # 搜索引擎
                            'rank_num': rank_num,  # 排名个数
                            'otherData':detail_obj_json
                        })
            if self.fugai_number == yiwancheng_obj:
                # 覆盖率  排名数量
                fugailv_sql = """select paiming_detail from fugai_Linshi_List where paiming_detail is not '0' and paiming_detail is not NULL ;"""
                objs = database_create_data.operDB(fugailv_sql, 'select')
                fugai_num = 0
                if len(objs['data']) > 0 and None not in objs['data'][0]:
                    for obj in objs['data']:
                        fugai_num += len(obj[0].split(','))
                        paiming_num += 1
                # 排名率
                if paiming_num and self.fugai_number:
                    if self.fugai_number and paiming_num != (0 and '0'):
                        paiminglv = int((int(paiming_num) / int(self.fugai_number)) * 100)
                # # 覆盖率
                if self.fugai_number:
                    if fugai_num != 0 and fugai_num:
                        fugailv = int((int(fugai_num) / int(self.fugai_number * 10)) * 100)
                whether_complete = True
        # print('覆盖当前已完成----------------------------------> ',yiwancheng_obj)
        exit_dict = {'data': data_list,
                     'total_data_num': self.fugai_number,  # 数据总数
                     'fugailv': fugailv,  # 覆盖率
                     'paiminglv': paiminglv,  # 排名率
                     'paiming_num': paiming_num,  # 排名数
                     'chongfu_num': self.fugai_chongfu_num,  # 重复数
                     'whether_complete': whether_complete,  # 全部完成 传True
                     'yiwancheng_obj': yiwancheng_obj  # 当前完成数量
                     }
        return json.dumps(exit_dict)

    # 覆盖查询 - 生成excel表格
    def set_fugai_save_select_result_value(self, data):
        if self.huoqu_fugai_time_stamp:
            wb = Workbook()
            ws = wb.active
            ws.title = '关键词覆盖查询'
            ws.cell(row=1, column=1, value="关键词覆盖查询")
            ws.cell(row=2, column=3, value="查询时间:")
            ws.cell(row=8, column=1, value="关键词")
            ws.cell(row=8, column=2, value="排名个数")
            ws.cell(row=8, column=3, value="排名情况")
            ws.cell(row=8, column=4, value="搜索引擎")
            ft1 = Font(name='宋体', size=22)
            a1 = ws['A1']
            a1.font = ft1

            # # 合并单元格        开始行      结束行       用哪列          占用哪列
            ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=4)
            ws.merge_cells(start_row=2, end_row=7, start_column=1, end_column=1)
            ws.merge_cells(start_row=2, end_row=7, start_column=2, end_column=2)
            ws.merge_cells(start_row=2, end_row=7, start_column=3, end_column=3)
            ws.merge_cells(start_row=2, end_row=7, start_column=4, end_column=4)
            # ws.merge_cells(start_row=2, end_row=5, start_column=5, end_column=5)

            # print('设置列宽')
            ws.column_dimensions['A'].width = 35
            ws.column_dimensions['B'].width = 13
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 30

            # # print('设置行高')
            ws.row_dimensions[1].height = 28

            # # print('文本居中')
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws['C2'].alignment = Alignment(horizontal='right', vertical='center')
            ws['A8'].alignment = Alignment(horizontal='center', vertical='center')
            ws['B8'].alignment = Alignment(horizontal='center', vertical='center')
            ws['C8'].alignment = Alignment(horizontal='center', vertical='center')
            ws['D8'].alignment = Alignment(horizontal='center', vertical='center')
            ws['D2'].alignment = Alignment(horizontal='left', vertical='center')

            ws2 = wb.create_sheet('sheet2')
            ws2.title = '关键词覆盖查询详情'
            ws2.cell(row=1, column=1, value="关键词覆盖查询详情")
            ws2.cell(row=2, column=4, value="查询时间:")
            ws2.cell(row=3, column=1, value="关键词")
            ws2.cell(row=3, column=2, value="名次")
            ws2.cell(row=3, column=3, value="标题")
            ws2.cell(row=3, column=4, value="链接")
            ws2.cell(row=3, column=5, value="规则")
            ws2.cell(row=3, column=6, value="搜索引擎")
            ft1 = Font(name='宋体', size=22)
            a1 = ws2['A1']
            a1.font = ft1

            # # 合并单元格        开始行      结束行       开始列          结束列
            ws2.merge_cells(start_row=1, end_row=1, start_column=1, end_column=6)
            ws2.merge_cells(start_row=2, end_row=2, start_column=4, end_column=5)

            # # print('设置列宽')
            ws2.column_dimensions['A'].width = 35
            ws2.column_dimensions['B'].width = 15
            ws2.column_dimensions['C'].width = 60
            ws2.column_dimensions['D'].width = 30
            ws2.column_dimensions['E'].width = 13
            ws2.column_dimensions['F'].width = 20
            #
            # # print('设置行高')
            ws2.row_dimensions[1].height = 80
            ws2.row_dimensions[2].height = 30
            #
            # # print('文本居中')
            ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws2['A3'].alignment = Alignment(horizontal='center', vertical='center')
            ws2['B3'].alignment = Alignment(horizontal='center', vertical='center')
            ws2['C3'].alignment = Alignment(horizontal='center', vertical='center')
            ws2['D3'].alignment = Alignment(horizontal='center', vertical='center')
            ws2['E3'].alignment = Alignment(horizontal='center', vertical='center')
            ws2['F3'].alignment = Alignment(horizontal='center', vertical='center')
            ws2['C2'].alignment = Alignment(horizontal='center', vertical='center')
            ws2['D2'].alignment = Alignment(horizontal='right', vertical='center')
            ws2['F2'].alignment = Alignment(horizontal='left', vertical='center')
            chaxun_time = self.huoqu_fugai_time_stamp  # 查询时间
            now_date = datetime.datetime.today().strftime('%Y-%m-%d %H-%M-%S')
            ws.cell(row=2, column=4, value="{chaxun_time}".format(chaxun_time=self.dangqian_chaxunfugai_time))
            ws2.cell(row=2, column=6, value="{chaxun_time}".format(chaxun_time=self.dangqian_chaxunfugai_time))
            sql = """select * from fugai_Linshi_List where time_stamp={time_stamp};""".format(
                time_stamp=chaxun_time)
            objs = database_create_data.operDB(sql, 'select')
            row = 9
            row_two = 4
            for obj in objs['data']:
                if str(obj[2]) == '1':
                    yinqing = '百度'
                elif  str(obj[2]) == '4':
                    yinqing = '手机百度'
                elif str(obj[2]) == '3':
                    yinqing = '360'
                elif str(obj[2]) == '6':
                    yinqing = '手机360'
                else:
                    yinqing = ''
                paiming_detail = obj[1].split(',')
                paming_num = 0
                for paiming in paiming_detail:
                    paming_num += 1
                ws.cell(row=row, column=1, value="{keyword}".format(keyword=obj[10]))
                ws.cell(row=row, column=2, value="{paming_num}".format(paming_num=paming_num))
                ws.cell(row=row, column=3, value="{paiming_detail}".format(paiming_detail=obj[1]))
                ws.cell(row=row, column=4, value="{search}".format(search=yinqing))
                row += 1
                if obj[8]:
                    json_detail = json.loads(obj[8])
                    for data_detail in json_detail:
                        ws2.cell(row=row_two, column=1, value="{keyword}".format(keyword=obj[10]))
                        ws2.cell(row=row_two, column=2, value="{paiming}".format(paiming=data_detail['rank']))
                        ws2.cell(row=row_two, column=3, value="{title}".format(title=data_detail['title']))
                        ws2.cell(row=row_two, column=4, value="{title_url}".format(title_url=data_detail['url']))
                        ws2.cell(row=row_two, column=5, value="{guize}".format(guize=data_detail['guize']))
                        ws2.cell(row=row_two, column=6, value="{search}".format(search=data_detail['search_engine']))
                        row_two += 1

            root = Tk()
            root.iconbitmap('./128.ico')
            root.withdraw()  # 隐藏
            dirname = askdirectory(parent=root, initialdir="/", title='选择导出路径 !')
            if dirname:
                if dirname == 'C:/':
                    tkinter.messagebox.showerror('错误', '请选择路径 !')
                    # tkinter.messagebox.showwarning('警告', '请选择路径 !')
                else:
                    task_name = '覆盖查询'
                    file_name = dirname.replace('\\', '/') + '/' + '{}.xlsx'.format(task_name + '_' + now_date)
                    wb.save(file_name)
                    tkinter.messagebox.showinfo('提示', '生成完毕 !')
            else:
                print('点击取消')
            root.destroy()  # 销毁

    # 查询覆盖 - 退出清空查询结果
    def set_fugai_delete_all_values(self, data):
        if data == 'delete_all_fugai':
            process_fugai = os.popen('taskkill /PID %s /F' % self.process_fugai.pid)
            subprocess.Popen(process_fugai, shell=True)
            delete_sql = """delete from fugai_Linshi_List;"""
            database_create_data.operDB(delete_sql, 'delete')

    # 退出停止进程
    def __del__(self):
        if 'win' in sys.platform:
            process = os.popen('taskkill /PID %s /F' % self.process.pid)
            subprocess.Popen(process, shell=True)
            process_shoulu = os.popen('taskkill /PID %s /F' % self.process_shoulu.pid)
            subprocess.Popen(process_shoulu, shell=True)
            process_fugai = os.popen('taskkill /PID %s /F' % self.process_fugai.pid)
            subprocess.Popen(process_fugai, shell=True)

    loginValue = pyqtProperty(str, fget=get_Loginvalue, fset=set_Loginvalue)
    # 重点词监护 - 增加任务
    createTaskListValue = pyqtProperty(str, fget=get_zhongdianci_create_task_list_value,
        fset=set_zhongdianci_create_value)
    # 重点词监护 - 查询修改前数据 传递id 查询该id任务列表
    updateTaskListValue = pyqtProperty(str, fget=get_zhongdianci_update_task_list_data_value,
        fset=set_zhongdianci_update_task_list_value)
    # 重点词监护 - 使用↑id 修改任务列表
    updateDatavalue = pyqtProperty(str, fget=zhanwei_zhushou, fset=set_zhongdianci_update_data_value)
    # 重点词监护 - 清空该id的 任务列表 详情数据
    emptyDataTaskDetail = pyqtProperty(str, fget=zhanwei_zhushou, fset=set_zhongdianci_select_id_task_detail_value)
    # 重点词监护 - 获取id 查询该id详情
    selectDataTaskDetail = pyqtProperty(str, fget=get_zhongdianci_select_id_select_task_detail_value,
        fset=set_zhongdianci_select_id_select_task_detail_value)
    # 重点词监护 - 立即监控 - 获取id 分辨 移动或pc端
    pcTaskValue = pyqtProperty(str, fget=zhanwei_zhushou, fset=set_pc_task_value)
    # 重点词监护 - 删除单个或多个 任务
    deleteDataTask = pyqtProperty(str, fget=zhanwei_zhushou, fset=delete_or_batch_delete_task)
    # 重点词监护 - 详情分页
    detailPage = pyqtProperty(str, fget=zhanwei_zhushou, fset=set_huoqu_page_detail_value)
    # 重点词监护 - 查看该任务所有详情 及关键词
    viewThesubTasksDetail = pyqtProperty(str, fget=get_view_The_subtasks_detail,
        fset=set_task_id_view_The_subtasks_task)
    # 重点词监护 - 保存查询结果 导出excl表格
    saveTheQueryResults = pyqtProperty(str, fget=zhanwei_zhushou, fset=set_save_select_results_task_excel_daochu)
    # 收录查询 - 筛选链接 查询 及 展示入库
    ShouLuChaXun = pyqtProperty(str, fget=get_shoulu_zhanshi_list_value, fset=set_shoulu_select_get_list_value)
    # 收录查询 - 详情分页查询
    shouluChaxunPage = pyqtProperty(str, fget=zhanwei_zhushou, fset=set_shoulu_chauxn_page_value)
    # 收录查询 - 导出excel表格
    setShouLuDaoChuExcel = pyqtProperty(str, fget=zhanwei_zhushou, fset=set_shoulu_save_select_result_value)
    # 收录查询 - 退出时删除所有收录数据
    shouluExitDeleteAll = pyqtProperty(str, fget=zhanwei_zhushou, fset=set_shoulu_delete_all_values)
    # 覆盖查询 - 筛选关键词 查询入库及展示
    fugaiChaXun = pyqtProperty(str, fget=get_fugai_zhanshi_list_value, fset=set_fugai_select_get_list_value)
    # 覆盖查询 - 分页查询
    fuGaiChaXunPage = pyqtProperty(str, fget=zhanwei_zhushou, fset=set_fugai_chaxun_xiangqing)
    # 覆盖查询 - 导出excel表格
    setFuGaiDaoChuExcel = pyqtProperty(str, fget=zhanwei_zhushou, fset=set_fugai_save_select_result_value)
    # 覆盖查询 - 退出时删除所有覆盖数据
    fugaiExitDeleteAll = pyqtProperty(str, fget=zhanwei_zhushou, fset=set_fugai_delete_all_values)


# PyQt 架构 与 数据库初始化类
class DaNao(object):
    def __init__(self):
        self.initDatabase = multiprocessing.Process(target=self.initDB)
        self.initDatabase.start()
        self.main_body()

    # 初始化数据库和表
    def initDB(self):
        if not os.path.exists(settings.db_file):
            conn = sqlite3.connect(settings.db_file)
            conn.cursor()
        if os.path.exists(settings.lock_file):
            os.remove(settings.lock_file)

        # 查询sqlite所有表  查询表是否存在 不在创建
        sql = """select name from sqlite_master  where type = 'table' order by name;"""
        objs = database_create_data.operDB(sql, 'select')
        sqlite_dbs = []
        if objs['data']:
            for obj in objs['data']:
                sqlite_dbs.append(obj[0])

        print('数据库的所有表名-=-=-==-=-=-=--=-===-=-=> ', sqlite_dbs)
        # 判断登录 信息表
        if 'Login_Message' not in sqlite_dbs:
            print('没有Login_Message表 ------- 创建Login_Message表')
            Login_Message = """
                  create table Login_Message (
                  id integer primary key autoincrement,
                  message text not null
                  )"""
            database_create_data.operDB(Login_Message, 'create')

        # 判断 任务列表
        if 'task_List' not in sqlite_dbs:
            print('没有Task_List表 ------- 创建Task_List表')

            """
            qiyong_status       启用状态
            task_name           任务名称
            task_jindu          查询任务的进度百分比
            task_start_time     任务查询时间
            task_status         任务状态
                                     1 未查询
                                     2 查询中
                                     3 已完成

            search_engine      需要查询的搜索引擎
                                   1 百度
                                   2 搜狗
                                   3 360
                                   4 手机百度
                                   5 手机搜狗
                                   6 手机360
                                   7 神马

            mohupipei          模糊匹配条件
            zhixing            是否正在查询
            next_datetime      下次查询时间
            keywords           查询的关键词

            """

            Task_List_sql = """CREATE TABLE task_List (
                 "id" INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT,
                 "qiyong_status" INTEGER,
                 "task_name" TEXT,
                 "task_jindu" INTEGER,
                 "task_start_time" TEXT,
                 "task_status" INTEGER DEFAULT 1,
                 "search_engine" TEXT,
                 "mohupipei" TEXT,
                 "zhixing" INTEGER,
                 "next_datetime" TEXT,
                 "keywords" TEXT
               );"""
            database_create_data.operDB(Task_List_sql, 'create')

        # 判断 任务详情
        if 'task_Detail' not in sqlite_dbs:
            print('没有Task_Detail表 ------- 创建Task_Detail表')

            """   id
                  tid INTEGER          归属任务
                  search_engine        搜索引擎
                  lianjie              条件链接
                  keywords             关键词
                  mohupipei            模糊匹配条件
                  create_time          创建时间
                  task_start_time      任务执行时间
                  time_stamp           时间戳
                  is_perform           是否执行
            """
            Task_Detail_sql = """CREATE TABLE task_Detail (
                  "id" INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT,
                  "tid" INTEGER,
                  "search_engine" TEXT,
                  "lianjie" TEXT,
                  "keywords" TEXT,
                  "mohupipei" TEXT,
                  "create_time" TEXT,
                  "task_start_time" TEXT,
                  "time_stamp" INTEGER,
                  "is_perform" TEXT,
                  CONSTRAINT "task_detail_tid" FOREIGN KEY ("tid") REFERENCES "Task_List" ("id") ON DELETE RESTRICT ON UPDATE RESTRICT
                );"""
            database_create_data.operDB(Task_Detail_sql, 'create')

        # 判断 关键词详情数据
        if 'task_Detail_Data' not in sqlite_dbs:
            print('没有task_Detail_Data ------- 创建task_Detail_Data')
            """
                id
                paiming              排名
                is_shoulu            收录
                tid                  父id详情表
                create_time          创建时间
            """
            task_Detail_Data_sql = """CREATE TABLE task_Detail_Data (
                  "id" integer NOT NULL PRIMARY KEY AUTOINCREMENT,
                  "paiming" integer,
                  "is_shoulu" integer,
                  "tid" integer,
                  "create_time" TEXT,
                  CONSTRAINT "task_Detail_Data_tid" FOREIGN KEY ("tid") REFERENCES "task_Detail" ("id") ON DELETE RESTRICT ON UPDATE RESTRICT
                );"""
            database_create_data.operDB(task_Detail_Data_sql, 'create')

        # 判断 收录临时 表
        if 'shoulu_Linshi_List' not in sqlite_dbs:
            print('没有shoulu_Linshi_List表  创建shoulu_Linshi_List表')
            """
                id
                url              链接
                is_shoulu        收录
                time_stamp       时间戳 为了更准确的展示出当前查询的所有数据
                title            创建时间
                search           引擎
                kuaizhao_time    百度快照
                status_code      状态码
                is_zhixing       是否执行
                shijianchuo      判断时间戳
            """
            shoulu_Linshi_List_sql = """CREATE TABLE shoulu_Linshi_List (
                    "id" INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT,
                    "url" TEXT,
                    "is_shoulu" TEXT,
                    "time_stamp" TEXT,
                    "title" TEXT,
                    "search" integer,
                    "kuaizhao_time" TEXT,
                    "status_code" integer,
                    "is_zhixing" text,
                    "shijianchuo" TEXT
                );"""
            database_create_data.operDB(shoulu_Linshi_List_sql, 'create')

        # 判断 覆盖临时 表
        if 'fugai_Linshi_List' not in sqlite_dbs:
            print('没有fugai_Linshi_List表 ,创建fugai_Linshi_List表')
            """
                本表区分父子级
                paiming_detail      排名详情 父级为全部排名 子级为单个排名
                search_engine       搜索引擎
                sousuo_guize        搜索规则
                time_stamp          时间戳  后台判断用
                chaxun_status       查询状态 为1 代表查询完成
                is_zhixing          是否执行
                shijianchuo         取数据判断
                json_detail_data    json详情数据
                zhanwei             占位
                keyword             关键词
            """
            fugai_Linshi_List_sql = """
                CREATE TABLE fugai_Linshi_List (
                  "id" INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT,
                  "paiming_detail" TEXT,
                  "search_engine" TEXT,
                  "sousuo_guize" TEXT,
                  "time_stamp" TEXT,
                  "chaxun_status" integer,
                  "is_zhixing" TEXT,
                  "shijianchuo" integer,
                  "json_detail_data" TEXT,
                  "zhanwei" integer,
                  "keyword" TEXT
                );
            """
            database_create_data.operDB(fugai_Linshi_List_sql, 'create')

        # 判断覆盖临时表详情
        # if 'fugai_Linshi_Detail_List' not in sqlite_dbs:
        #     print('没有fugai_Linshi_Detail_List表 ,创建fugai_Linshi_Detail_List表')
        #     """
        #       paiming_detail   排名详情
        #       search           搜索引擎
        #       title            标题
        #       guize            规则
        #       title_url        标题链接
        #       tid              父id
        #       keyword          关键词
        #     """
        #     fugai_Linshi_Detail_List_sql = """CREATE TABLE fugai_Linshi_Detail_List (
        #       "id" INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT,
        #       "paiming_detail" integer,
        #       "search" TEXT,
        #       "title" TEXT,
        #       "guize" TEXT,
        #       "title_url" TEXT,
        #       "tid" INTEGER,
        #       "keyword" TEXT,
        #       CONSTRAINT "shoulu_linshi_list_tid" FOREIGN KEY ("tid") REFERENCES "fugai_Linshi_List" ("id") ON DELETE RESTRICT ON UPDATE RESTRICT
        #     );"""
        #     database_create_data.operDB(fugai_Linshi_Detail_List_sql, 'create')

        initDatabase = os.popen('taskkill /PID %s /F' % self.initDatabase.pid)
        subprocess.Popen(initDatabase, shell=True)

    # 主体 函数
    def main_body(self):
        win32_width = win32api.GetSystemMetrics(0) * 0.8
        win32_height = win32api.GetSystemMetrics(1) * 0.8

        app = QApplication(sys.argv)
        win = QWidget()
        win.resize(win32_width, win32_height)
        win.setWindowTitle('诸葛大脑')
        # app.setWindowIcon(QIcon(os.path.join(os.getcwd(), '128.ico')))
        app.setWindowIcon(QIcon('./128.ico'))

        # 创建垂直布局
        layout = QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        win.setLayout(layout)
        view = QWebEngineView()
        # view.load(QUrl('http://wenda.zhugeyingxiao.com'))
        # view.load(QUrl('http://192.168.10.240:8081'))
        view.load(QUrl('http://zhugedanao1.zhugeyingxiao.com'))

        # 简单理解就是将这个控件(QWidget)的几何内容(宽高位置等)，赋值给qr
        # qr = view.frameGeometry()
        # 计算出你的显示器的屏幕分辨率。根据得到的分辨率我们得到屏幕的中心点。
        # cp = QDesktopWidget().availableGeometry().center()
        # 我们的矩形(qr)已有宽度和高度，现在设置移动矩形的中心(moveCenter)到屏幕的中心点(cp)，矩形的尺寸是不变的。
        # qr.moveCenter(cp)
        # 移动应用程序窗口的左上角到qr矩形的左上角，从而使应用程序窗口显示在屏幕的中心
        # view.move(qr.topLeft())
        # 创建一个 QWebChannel 对象, 用来传递 PyQt的参数到 Js
        channel = QWebChannel()
        myObj = Danao_Inter_Action()
        channel.registerObject("bridge", myObj)
        view.page().setWebChannel(channel)

        # 把 QWebEngineView 控件加载到 layout 布局中
        layout.addWidget(view, 0)

        # 显示窗口和运行
        # win.showMaximized() # 窗口最大化
        win.show()          # 普通窗口

        sys.exit(app.exec_())


if __name__ == '__main__':
    multiprocessing.freeze_support()
    obj = DaNao()


